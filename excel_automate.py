"""
just working through automate the boring stuff
"""
from openpyxl import Workbook

wb = Workbook()

# grab the active worksheet
ws = wb.active
ws.title = "Sheet1"  # trying to create the first sheet

# filling in values based on the site. Starting with cell A1 - A7
ws["A1"] = "4/5/2015 1:34:02 PM"
ws["A2"] = "4/5/2015 3:41:23 AM"
ws["A3"] = "4/6/2015 12:46:51 PM"
ws["A4"] = "4/8/2015 8:59:43 AM"
ws["A5"] = "4/10/2015 2:07:00 AM"
ws["A6"] = "4/10/2015 6:10:37 PM"
ws["A7"] = "4/10/2015 2:40:46 AM"

# filling in values for B1 - B7
ws["B1"] = "Apples"
ws["B2"] = "Cherries"
ws["B3"] = "Pears"
ws["B4"] = "Oranges"
ws["B5"] = "Apples"
ws["B6"] = "Bananas"
ws["B7"] = "Strawberries"

# filling in values for C1 - C7
ws["C1"] = "73"
ws["C2"] = "85"
ws["C3"] = "14"
ws["C4"] = "52"
ws["C5"] = "152"
ws["C6"] = "23"
ws["C7"] = "98"

wb.save("example.xlsx")

sheet = wb.get_sheet_by_name("Sheet1")
print(sheet["A1"])
print(sheet["A1"].value)
c = sheet["B1"]
print(c.value)
print("Row {}, Column {}, is {}".format(str(c.row), c.column, c.value))
print("Cell {} is {}".format(c.coordinate, c.value))
print(sheet["C1"].value)
print("Selected cell: {}".format(sheet.cell(row=1, column=2)))
print("Sheet1.B1 value: {}".format(sheet.cell(row=1, column=2).value))

for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

print("Sheet Max Row: {}".format(sheet.max_row))
print("Sheet Max Column: {}".format(sheet.max_column))
